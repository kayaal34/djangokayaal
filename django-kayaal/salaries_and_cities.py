import csv
from operator import itemgetter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


class report:
    def __init__(self, vacancies_count_by_years, vacancies_count_by_years_for_profession, salary_by_years,
                 salary_by_years_for_profession, vacancies_count_by_cities, vacancies_share_by_cities, salary_by_cities,
                 profession):
        self.vacancies_count_by_years = vacancies_count_by_years
        self.vacancies_count_by_years_for_profession = vacancies_count_by_years_for_profession
        self.salary_by_years = salary_by_years
        self.salary_by_years_for_profession = salary_by_years_for_profession
        self.vacancies_count_by_cities = vacancies_count_by_cities
        self.vacancies_share_by_cities = vacancies_share_by_cities
        self.salary_by_cities = salary_by_cities
        self.side = Side(style='thin', color="000000")
        self.thin_border = Border(left=self.side, right=self.side, top=self.side, bottom=self.side)
        self.profession = "frontend"

    def make_column_stylied(self, ws, index_column, max_row_index):
        max_value = -6
        for row in range(1, max_row_index):
            cell = ws.cell(column=index_column, row=row)
            if max_value < len(str(cell.value)):
                max_value = len(str(cell.value))
            cell.border = self.thin_border
        ws.column_dimensions[get_column_letter(index_column)].width = max_value + 2

    def fill_column_by_years(self, ws, column_index, dictionary_values, min_year, max_year):
        max_row_index = max_year - min_year
        for row, value in zip([index for index in range(2, max_row_index + 2)],
                              [dictionary_values[value] for value in range(min_year, max_year)]):
            ws.cell(row=row, column=column_index, value=value)

    def fill_column_by_cities(self, ws, column_index, max_row_index, dictionary_values, keys):
        for row, value in zip([index for index in range(2, max_row_index + 2)],
                              [dictionary_values[key] for key in keys]):
            ws.cell(row=row, column=column_index, value=value)

    def make_ws_by_years(self, ws):
        min_year, max_year = 2004, 2022
        max_column_index = 5
        max_row_index = max_year - min_year
        ws.title = "Статистика по годам"
        self.make_titles(max_column_index,
                         ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий",
                          f"Количество вакансий - {self.profession}"],
                         ws)
        self.fill_column_by_years(ws, 1,
                                  {x: y for x, y in zip(range(min_year, max_year + 1), range(min_year, max_year + 1))},
                                  min_year, max_year + 1)
        self.fill_column_by_years(ws, 2, self.salary_by_years, min_year, max_year + 1)
        self.fill_column_by_years(ws, 3, self.salary_by_years_for_profession, min_year, max_year + 1)
        self.fill_column_by_years(ws, 4, self.vacancies_count_by_years, min_year, max_year + 1)
        self.fill_column_by_years(ws, 5, self.vacancies_count_by_years_for_profession, min_year, max_year + 1)
        for i in range(1, 6):
            self.make_column_stylied(ws, i, max_row_index + 3)

    def make_ws_by_cities(self, ws):
        max_column_index = 5
        max_row_index = 11
        ws.title = "Статистика по городам"
        self.make_titles(max_column_index, ["Город", "Уроввень зарплат", "", "Город", "Доля вакансий"], ws)
        keys = self.salary_by_cities.keys()
        self.fill_column_by_cities(ws, 1, max_row_index, {x: y for x, y in zip(keys, keys)}, keys)
        self.fill_column_by_cities(ws, 2, max_row_index, self.salary_by_cities, keys)
        keys = self.vacancies_share_by_cities.keys()
        self.fill_column_by_cities(ws, 4, max_row_index, {x: y for x, y in zip(keys, keys)}, keys)
        self.fill_column_by_cities(ws, 5, max_row_index, self.vacancies_share_by_cities, keys)
        for i in range(1, 6):
            self.make_column_stylied(ws, i, max_row_index + 1)
        ws.column_dimensions[get_column_letter(3)].width = 2
        for cell in ws['E']:
            cell.number_format = '0.00%'

    def make_titles(self, max_column_index, names, ws):
        for cols_index, name in zip([i for i in range(1, max_column_index + 1)], names):
            ws.cell(row=1, column=cols_index, value=name).font = Font(bold=True)

    def generate_excel(self):
        wb = Workbook()
        ws_by_years = wb.active
        ws_by_cities = wb.create_sheet()
        self.make_ws_by_years(ws_by_years)
        self.make_ws_by_cities(ws_by_cities)
        wb.save("report.xlsx")
        return ws_by_years, ws_by_cities

    def generate_normal_bar_graph(self, title, first_line_legend, second_line_legend, fisrt_dict, second_dict, ax):
        labels = [i for i in range(2004, 2023)]
        salary_by_year = [fisrt_dict[key] for key in range(2004, 2023)]
        salary_by_year_for_profession = [second_dict[key] for key in range(2004, 2023)]
        x = np.arange(len(labels))
        width = 0.35
        ax.bar(x - width / 2, salary_by_year, width, label=first_line_legend)
        ax.bar(x + width / 2, salary_by_year_for_profession, width, label=second_line_legend)
        ax.set_title(title)
        ax.set_xticks(x, labels, rotation=90)
        ax.legend(fontsize=8)
        ax.yaxis.grid(True)

    def generate_reverse_bar_graph(self, ax):
        cities = self.salary_by_cities.keys()
        y_pos = np.arange(len(cities))
        performance = [self.salary_by_cities[key] for key in cities]
        ax.barh(y_pos, performance, align='center')
        ax.set_yticks(y_pos, labels=cities)
        ax.invert_yaxis()
        ax.set_title('Уровень зарплат по городам')
        ax.xaxis.grid(True)

    def generate_pie(self, ax):
        keys = [*self.vacancies_count_by_cities]
        top_keys, other_keys = keys[:10], keys[10:]
        x = [self.vacancies_count_by_cities[key] for key in top_keys]
        x.append(sum([self.vacancies_count_by_cities[key] for key in other_keys]))
        top_keys.append("Другие")
        ax.pie(x)
        ax.legend(loc="lower right", labels=top_keys, fontsize=6)
        ax.set_title('Доля вакансий по городам')

    def generate_image(self):
        fig = plt.figure()
        plt.rc('xtick', labelsize=8)
        plt.rc('ytick', labelsize=8)
        ax1 = fig.add_subplot(221)
        ax2 = fig.add_subplot(222)
        ax3 = fig.add_subplot(223)
        ax4 = fig.add_subplot(224)
        self.generate_normal_bar_graph("Уровень зарплат по годам", "средняя з/п", f"з/п {self.profession}",
                                       self.salary_by_years, self.salary_by_years_for_profession, ax1)
        self.generate_normal_bar_graph("Количество вакансий по годам", "Количество вакансий",
                                       f"Количество вакансий {self.profession}",
                                       self.vacancies_count_by_years, self.vacancies_count_by_years_for_profession, ax2)
        self.generate_reverse_bar_graph(ax3)
        self.generate_pie(ax4)
        plt.tight_layout()
        fig.set_size_inches(10, 7)
        plt.savefig('graph.png')

    def remake_to_percantage(self, ws, column_index):
        for row in range(2, ws.max_row + 1):
            ws.cell(column=column_index, row=row).value = str(round((ws.cell(column=column_index, row=row).value * 100), 2)).replace(".", ",") + "%"
        return ws

    def generate_pdf(self):
        self.generate_image()
        year_stat, cities_stat = self.generate_excel()
        cities_stat = self.remake_to_percantage(cities_stat, 5)
        image = "graph.png"
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(profession=self.profession, image=image, year_stat=year_stat, cities_stat=cities_stat)
        config = pdfkit.configuration(wkhtmltopdf=r"C:\wkhtmltopdf\bin\wkhtmltopdf.exe")
        pdfkit.from_string(pdf_template, "report.pdf", configuration=config, options={'enable-local-file-access': None})


class Vacancy:
    def __init__(self, dictionary):
        self.name = dictionary["name"]
        self.area_name = dictionary["area_name"]
        self.published_at = int(dictionary["published_at"][:4])
        self.currency_to_rub = {"AZN": 35.68,
                                "BYR": 23.91,
                                "EUR": 59.90,
                                "GEL": 21.74,
                                "KGS": 0.76,
                                "KZT": 0.13,
                                "RUR": 1,
                                "UAH": 1.64,
                                "USD": 60.66,
                                "UZS": 0.0055, }
        if dictionary["salary_from"] == "" and dictionary["salary_to"] == "" and dictionary["salary_currency"] == "":
            self.salary = 0
        elif dictionary["salary_from"] == "":
            self.salary = float(dictionary["salary_to"]) * self.currency_to_rub[dictionary["salary_currency"]]
        elif dictionary["salary_to"] == "":
            self.salary = float(dictionary["salary_from"]) * self.currency_to_rub[dictionary["salary_currency"]]
        else:
            self.salary = (float(dictionary["salary_from"]) + float(dictionary["salary_to"])) / 2 * self.currency_to_rub[
                dictionary["salary_currency"]]

    def __str__(self):
        return self.name + self.published_at




class DataSet:
    def __init__(self, file_name, profession):
        self.profession = profession
        self.file_name = file_name
        csv_read = self.csv_reader()
        dictionaries = self.csv_filter(csv_read[1], csv_read[0])
        vacancies_list = []
        for dictionary in dictionaries:
            vacancies_list.append(Vacancy(dictionary))
        self.vacancies_elements = vacancies_list
        self.count_year_vacancies = self.count_vacancies_by_years()
        self.count_year_vacancies_for_profession = self.count_profession_vacancies_by_years()
        self.years_salary = self.get_years_salary()
        self.years_salary_for_profession = self.get_profession_years_salary()
        self.count_cities_vacancies = self.count_vacancies_by_cities()
        self.vacancies_share_by_cities = self.get_vacancies_share_by_cities()
        self.cities_salary = self.get_cities_salary()

    def print_information(self):
        print("Динамика уровня зарплат по годам: " + str(self.years_salary))
        print("Динамика количества вакансий по годам: " + str(self.count_year_vacancies))
        print("Динамика уровня зарплат по годам для выбранной профессии: " + str(
            self.years_salary_for_profession))
        print("Динамика количества вакансий по годам для выбранной профессии: " + str(
            self.count_year_vacancies_for_profession))
        print("Уровень зарплат по городам (в порядке убывания): " + str(self.cities_salary))
        print("Доля вакансий по городам (в порядке убывания): " + str(self.vacancies_share_by_cities))

    def check_rows_count(self, rows_count):
        if rows_count == 0:
            print("Пустой файл")
            exit()
        if rows_count == 1:
            print("Нет данных")
            exit()

    def count_profession_vacancies_by_years(self):
        dictionary = {}
        for vacancy in self.vacancies_elements:
            name = vacancy.name.lower().split()
            flag = True
            for element in name:
                if element in self.profession:
                    flag = False
                    break
            if flag:
                continue
            dictionary[vacancy.published_at] = (
                dictionary[vacancy.published_at] + 1 if vacancy.published_at in dictionary else 1)
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        if len(dictionary) == 0:
            dictionary[2022] = 0
        return dictionary

    def get_years_salary(self):
        dictionary = {}
        for vacancy in self.vacancies_elements:
            dictionary[vacancy.published_at] = (
                dictionary[vacancy.published_at] + vacancy.salary if vacancy.published_at in dictionary else vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.count_year_vacancies[key])
        return dict(sorted(dictionary.items(), key=itemgetter(0)))

    def get_ten_items(self, dictionary):
        result_dictionary = {}
        for key, i in zip(dictionary, [i for i in range(10)]):
            result_dictionary[key] = round(dictionary[key], 4)
        return result_dictionary

    def csv_reader(self):
        vacancies, headlines = [], []
        length, rows_count = 0, 0
        first = True
        with open(self.file_name, encoding="utf-8-sig") as File:
            file = csv.reader(File)
            for row in file:
                rows_count += 1
                if first:
                    headlines = row
                    first = False
                else:
                    vacancies.append(row)
        self.check_rows_count(rows_count)
        return headlines, vacancies

    def get_vacancies_share_by_cities(self):
        dictionary = {}
        for key in self.count_cities_vacancies:
            if self.count_cities_vacancies[key] / len(self.vacancies_elements) >= 0.01:
                dictionary[key] = self.count_cities_vacancies[key] / len(self.vacancies_elements)
        return self.get_ten_items(dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True)))

    def get_cities_salary(self):
        dictionary = {}
        for vacancy in self.vacancies_elements:
            if self.count_cities_vacancies[vacancy.area_name] / len(self.vacancies_elements) < 0.01:
                continue
            dictionary[vacancy.area_name] = (
                dictionary[
                    vacancy.area_name] + vacancy.salary if vacancy.area_name in dictionary else vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.count_cities_vacancies[key])
        return self.get_ten_items(dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True)))

    def csv_filter(self, reader, list_naming):
        dictionaries = []
        for vacancy in reader:
            dictionary = {}
            for name, item in zip(list_naming, vacancy):
                dictionary[name] = item
            dictionaries.append(dictionary)
        return dictionaries

    def count_vacancies_by_years(self):
        dictionary = {}
        for vacancy in self.vacancies_elements:
            dictionary[vacancy.published_at] = (
                dictionary[vacancy.published_at] + 1 if vacancy.published_at in dictionary else 1)
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        return dictionary

    def get_profession_years_salary(self):
        dictionary = {}
        for vacancy in self.vacancies_elements:
            name = vacancy.name.lower().split()
            flag = True
            for element in name:
                if element in self.profession:
                    flag = False
                    break
            if flag:
                continue
            dictionary[vacancy.published_at] = (
                dictionary[
                    vacancy.published_at] + vacancy.salary if vacancy.published_at in dictionary else vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.count_year_vacancies_for_profession[key])
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        if len(dictionary) == 0:
            dictionary[2022] = 0
        return dictionary

    def count_vacancies_by_cities(self):
        dictionary = {}
        for vacancy in self.vacancies_elements:
            dictionary[vacancy.area_name] = (
                dictionary[
                    vacancy.area_name] + 1 if vacancy.area_name in dictionary else 1)
        return dictionary


file_name = "vacancies_with_skills.csv"
profession = ['frontend', 'фронтенд', 'вёрстка', 'верстка', 'верста', 'front end', 'angular', 'html', 'css', 'react', 'vue']
dataset = DataSet(file_name, profession)
report(dataset.count_year_vacancies, dataset.count_year_vacancies_for_profession, dataset.years_salary,
       dataset.years_salary_for_profession,
       dataset.count_cities_vacancies, dataset.vacancies_share_by_cities, dataset.cities_salary,
       dataset.profession).generate_pdf()
